# phase-4.py
import os
import sys
import pandas as pd
import numpy as np
import warnings
from datetime import time, datetime
import logging

# Suppress pandas datetime parsing warnings
warnings.filterwarnings('ignore', message='Could not infer format')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
# ===============================
# CONFIG
# ===============================
PHASE3_FILE = "phase-3results/Phase3_results.xlsx"
ONE_MIN_DATA_DIR = "downloaded_data/1min/1min"   # 1-minute candles
OUTPUT_DIR = "phase-4results"
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"phase4_backtest_1m_{ts}.xlsx")
LOG_FILE = os.path.join(OUTPUT_DIR, f"phase4_backtest_1m_{ts}.log")

# Capital & Risk Config
import config_manager
P4_CFG = config_manager.get_phase_config("phase4")

# Capital & Risk Config
C_PER_DAY = P4_CFG.get("C_PER_DAY", 1000000)
L_PCT = P4_CFG.get("L_PCT", 0.02)
C_PCT = P4_CFG.get("C_PCT", 0.50)

# Phase-4C Config
S_PCT = 0.02  # 2% stop-loss (TEST 2)
R_MULT = 1.0

MIN_TRADE_VALUE = 5000
FORCE_EXIT_TIME = config_manager.get_time_from_config(P4_CFG, "FORCE_EXIT_TIME") or time(15, 10)

# Transaction cost and slippage
TRANSACTION_COST_PCT = P4_CFG.get("TRANSACTION_COST_PCT", 0.0005)
SLIPPAGE_PCT = P4_CFG.get("SLIPPAGE_PCT", 0.001)

# Minimum edge filter
COST_BUFFER = 0.0025

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ===============================
# LOGGING
# ===============================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, mode="a", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("phase4")

# ===============================
# TRAILING STOP CALCULATION
# ===============================
def calculate_trailing_stop(entry_price, initial_stop, current_high, current_stop):
    """
    Calculate trailing stop-loss based on 3-tier R-multiple strategy:
    - 1R: Move to breakeven + costs
    - 1.5R: Lock 0.5R profit
    - 2R+: Trail by 1R distance from high
    
    Args:
        entry_price: Entry price of the trade
        initial_stop: Initial stop-loss price
        current_high: Highest price achieved since entry
        current_stop: Current stop-loss price
    
    Returns:
        new_stop: Updated stop-loss price
    """
    R = entry_price - initial_stop
    
    # Cost buffer (approx 0.1% or fixed amount to cover brokerage)
    cost_buffer = entry_price * 0.001 
    
    new_stop = current_stop
    
    # 1. Check Breakeven Trigger (1R)
    if current_high >= (entry_price + R):
        # Move to Entry + Costs (Ensure we don't move stop DOWN if it's already higher)
        new_stop = max(new_stop, entry_price + cost_buffer)
        
    # 2. Check Profit Locking Trigger (1.5R)
    if current_high >= (entry_price + (1.5 * R)):
        # Lock 0.5R profit
        new_stop = max(new_stop, entry_price + (0.5 * R))
        
    # 3. Dynamic Trail (Deep Profit > 2R)
    if current_high >= (entry_price + (2.0 * R)):
        # Trail by 1R distance from High
        new_stop = max(new_stop, current_high - R)
        
    return new_stop

# ===============================
# LOAD PHASE-3 OUTPUT
# ===============================
df = pd.read_excel(PHASE3_FILE)

df = df.rename(columns={
    "Stock": "symbol",
    "Entry Price (₹)": "entry_price",
    "Stop-Loss (₹)": "stop_loss",  # From Phase-3
    "Target (₹)": "target",        # From Phase-3
    "Date": "date",
    "Entry Time": "buy_time",
    "Entry Mode": "mode"  # A, B, or C from Phase-3
})

# Read VelocityScore if available, otherwise use default
if "VelocityScore" in df.columns:
    df["velocity_score"] = pd.to_numeric(df["VelocityScore"], errors="coerce").fillna(75.0)
else:
    # Default VelocityScore if not provided by Phase-3
    # TODO: Phase-3 should output VelocityScore in future
    df["velocity_score"] = 75.0

required_cols = {"symbol", "entry_price", "stop_loss", "target", "date", "buy_time"}
# Prepare output: multi-sheet Excel with formatting
missing = required_cols - set(df.columns)
if missing:
    raise ValueError(f"Missing required columns: {missing}")

# ===============================
# MINIMUM EDGE FILTER
# ===============================
# Filter: (Target - Entry) / Entry >= CostBuffer
# Prevents "death by a thousand paper cuts"
initial_count = len(df)
df["potential_profit_pct"] = ((df["target"] - df["entry_price"]) / df["entry_price"])
df = df[df["potential_profit_pct"] >= COST_BUFFER].copy()
filtered_count = initial_count - len(df)
logger.info("Minimum edge filter: Removed %s trades with <%.2f%% potential profit", filtered_count, COST_BUFFER*100)
logger.info("Remaining trades: %s", len(df))
df = df.drop(columns=["potential_profit_pct"])  # Clean up temporary column

# ===============================
# PHASE-4A — WEIGHTS (VelocityScore-based)
# ===============================
# Formula: s_i = VelocityScore - 50 (clip at 0)
#          w_i = s_i / Σ(s_i)

df["s_i"] = df["velocity_score"] - 50
df["s_i"] = df["s_i"].clip(lower=0)  # If < 0, set to 0

# ===============================
# PHASE-4B — DAILY CAPITAL ALLOCATION
# ===============================
# Allocate ₹1,00,000 per day (not total)
# Group by date and calculate weights within each day

all_days_data = []

for trade_date, day_df in df.groupby("date"):
    day_df = day_df.copy()
    
    # Calculate weights for this day only
    sum_s_i = day_df["s_i"].sum()
    if sum_s_i > 0:
        day_df["weight"] = day_df["s_i"] / sum_s_i
    else:
        day_df["weight"] = 0.0
    
    # Risk allocation for this day
    Lcap_day = C_PER_DAY * L_PCT
    day_df["loss_cap"] = Lcap_day * day_df["weight"]
    
    # ===============================
    # PHASE-4C — USE PHASE-3 PRICE LEVELS
    # ===============================
    # Use stop-loss and target from Phase-3 (don't recalculate!)
    day_df["stop_price"] = pd.to_numeric(day_df["stop_loss"], errors="coerce")
    day_df["target_price"] = pd.to_numeric(day_df["target"], errors="coerce")
    day_df["risk_per_share"] = day_df["entry_price"] - day_df["stop_price"]
    
    # ===============================
    # PHASE-4D — QUANTITY (per day)
    # ===============================
    day_df["risk_per_share"] = pd.to_numeric(day_df["risk_per_share"], errors="coerce").fillna(0.0)
    day_df["loss_cap"] = pd.to_numeric(day_df["loss_cap"], errors="coerce").fillna(0.0)
    day_df["entry_price"] = pd.to_numeric(day_df["entry_price"], errors="coerce").fillna(0.0)
    
    def safe_floor_div(a, b):
        try:
            a = float(a)
            b = float(b)
        except Exception:
            return 0.0
        if b <= 0:
            return 0.0
        return np.floor(a / b)
    
    day_df["qty_risk"] = day_df.apply(lambda r: safe_floor_div(r["loss_cap"], r["risk_per_share"]), axis=1)
    day_df["qty_cap"] = np.floor((C_PER_DAY * C_PCT) / day_df["entry_price"]).replace([np.inf, -np.inf], 0).fillna(0)
    day_df["quantity"] = np.minimum(day_df["qty_risk"], day_df["qty_cap"])
    
    day_df["trade_value"] = day_df["quantity"] * day_df["entry_price"]
    
    # ===============================
    # PHASE-4E — PORTFOLIO SCALING (per day)
    # ===============================
    total_deployed_day = day_df["trade_value"].sum()
    if total_deployed_day > C_PER_DAY:
        alpha = C_PER_DAY / total_deployed_day
        day_df["quantity"] = np.floor(day_df["quantity"] * alpha)
        day_df["trade_value"] = day_df["quantity"] * day_df["entry_price"]
    
    # Remove zero-quantity rows for this day
    day_df = day_df[day_df["quantity"] > 0].copy()
    
    all_days_data.append(day_df)

# Combine all days
df = pd.concat(all_days_data, ignore_index=True) if all_days_data else pd.DataFrame()

# ===============================
# PHASE-4 EXIT RESOLUTION (5-MIN BACKTEST)
# ===============================
sell_prices = []
sell_times = []
exit_reasons = []
pnls = []

exec_entry_times = []
exec_entry_prices = []

# State persistence lists for Phase-4B continuity
state_highest_prices = []
state_current_stops = []
state_last_candle_times = []
state_trailing_armed = [] # FIX #1: Persist trailing armed state

df = df.copy()

for _, row in df.iterrows():
    symbol = row["symbol"]
    buy_price = row["entry_price"]  # default
    stop = row["stop_price"]
    target = row["target_price"]
    qty = row["quantity"]
    buy_time = row["buy_time"]

    # normalize buy_time to a time object if possible
    if pd.isna(buy_time):
        buy_time = time(0, 0)
    elif isinstance(buy_time, str):
        try:
            bt = pd.to_datetime(buy_time).time()
            buy_time = bt
        except Exception:
            try:
                parts = [int(x) for x in buy_time.split(":")]
                buy_time = time(*parts[:3])
            except Exception:
                buy_time = time(0, 0)
    elif isinstance(buy_time, pd.Timestamp):
        buy_time = buy_time.time()

    # Build path to 1-minute CSV: downloaded_data/1min_candles/<SYMBOL>.csv
    date_val = row.get("date")
    date_str = None
    if pd.isna(date_val):
        date_str = None
    else:
        try:
            date_ts = pd.to_datetime(date_val)
            date_str = date_ts.strftime("%Y-%m-%d")
        except Exception:
            date_str = str(date_val)

    # Updated Logic: Data is stored as downloaded_data/1min/1min/<SYMBOL>/<DATE>.csv
    one_min_file = None
    if date_str:
        one_min_file = os.path.join(ONE_MIN_DATA_DIR, symbol, f"{date_str}.csv")
    
    # Fallback to old path if not found (optional, but good for safety)
    if not one_min_file or not os.path.exists(one_min_file):
        # check if maybe it's the old structure
        alt_file = os.path.join(ONE_MIN_DATA_DIR, f"{symbol}.csv")
        if os.path.exists(alt_file):
            one_min_file = alt_file


    sell_price = buy_price
    sell_time = None
    exit_reason = "NO_EXIT"
    had_post_buy_candle = False
    last_candle_close = None
    last_candle_time = None
    
    # Initialize trailing stop variables
    initial_stop = stop  # Store the original stop for R calculation
    current_stop = stop  # This will be updated dynamically
    highest_price = buy_price  # Track the highest price achieved
    trailing_armed = False     # Track if 1R target was crossed

    if one_min_file and os.path.exists(one_min_file):
        mdf = pd.read_csv(one_min_file)
        # normalize column names to lowercase
        mdf.columns = [c.lower() for c in mdf.columns]

        # 1-min data has 'date' column with full datetime (e.g., '2025-11-03 09:15:00+05:30')
        if "date" in mdf.columns:
            try:
                mdf["datetime"] = pd.to_datetime(mdf["date"])
            except Exception:
                mdf["datetime"] = pd.NaT
        elif "time" in mdf.columns and date_str:
             # If only 'time' exists (e.g. 09:15), combine with the trade date
             try:
                 mdf["datetime"] = pd.to_datetime(f"{date_str} " + mdf["time"])
             except Exception:
                 mdf["datetime"] = pd.NaT
        else:
            # cannot interpret timestamps; skip
            mdf["datetime"] = pd.NaT

        # Filter for the specific trade date only
        if date_str:
            mdf = mdf[mdf["datetime"].dt.strftime("%Y-%m-%d") == date_str]
        
        mdf = mdf.sort_values("datetime")

        # -------------------------------
        # MODEL A1 ENTRY EXECUTION (BROKER-REALISTIC)
        # Phase-3 confirms on 5m close
        # Execute at OPEN of the NEXT 1-min candle
        # -------------------------------

        actual_entry_time = None
        actual_entry_price = None

        for _, m in mdf.iterrows():
            if pd.isna(m.get("datetime")):
                continue

            t = m["datetime"].time()

            # first 1-minute candle AFTER Phase-3 confirmation
            if t > buy_time:
                actual_entry_time = t
                actual_entry_price = float(m.get("open", buy_price))
                break

        # Apply execution
        if actual_entry_time:
            buy_time = actual_entry_time
            buy_price = actual_entry_price

        for _, m in mdf.iterrows():
            # skip rows without valid datetime
            if pd.isna(m.get("datetime")):
                continue
            t = m["datetime"].time()

            # start checking only from the candle strictly after the buy_time
            if t <= buy_time:
                continue

            had_post_buy_candle = True
            last_candle_close = float(m.get("close", np.nan)) if not pd.isna(m.get("close", np.nan)) else None
            last_candle_time = t

            high = float(m.get("high", np.nan))
            low = float(m.get("low", np.nan))
            op = float(m.get("open", np.nan))
            close = float(m.get("close", np.nan))

            # Update highest price if this candle made a new high
            if not np.isnan(high) and high > highest_price:
                highest_price = high
                # Check 1R trigger to arm trailing stop
                if not trailing_armed and highest_price >= (buy_price + (buy_price - initial_stop)):
                    trailing_armed = True
                
                # Recalculate trailing stop based on new high
                current_stop = calculate_trailing_stop(buy_price, initial_stop, highest_price, current_stop)

            # Check exits using the DYNAMIC trailing stop
            hit_stop = (not np.isnan(low)) and (low <= current_stop)
            hit_target = (not np.isnan(high)) and (high >= target)

            # Both hit in same candle → approximate intrabar order using distance from open
            if hit_stop and hit_target:
                dist_stop = abs(op - current_stop) if not np.isnan(op) else float('inf')
                dist_target = abs(op - target) if not np.isnan(op) else float('inf')
                if dist_stop <= dist_target:
                    sell_price = current_stop
                    exit_reason = "TRAILING_STOP_INTRABAR"
                else:
                    sell_price = target
                    exit_reason = "TARGET_INTRABAR"
                sell_time = t
                break

            if hit_stop:
                sell_price = current_stop
                sell_time = t
                # Determine which tier of trailing stop was hit
                if current_stop > buy_price:
                    exit_reason = "TRAILING_STOP_PROFIT"
                else:
                    exit_reason = "STOP_LOSS"
                break

            if hit_target:
                sell_price = target
                sell_time = t
                exit_reason = "TARGET"
                break

            # Force-exit at or after FORCE_EXIT_TIME using candle close
            if t >= FORCE_EXIT_TIME:
                sell_price = close if not np.isnan(close) else m.get("close", buy_price)
                sell_time = t
                exit_reason = "TIME_EXIT_1510"
                break

        # If no explicit exit was found but there were candles after the buy_time,
        # close at the last available candle close to approximate end-of-data exit.
        if exit_reason == "NO_EXIT" and had_post_buy_candle and last_candle_close is not None:
            sell_price = last_candle_close
            sell_time = last_candle_time
            exit_reason = "NO_EXIT_LASTCANDLE"
    else:
        # No 1-minute data available for this symbol — close at buy price and mark
        sell_price = buy_price
        sell_time = buy_time
        exit_reason = "NO_1M_DATA"
        # Set default state for trades with no data
        highest_price = buy_price
        current_stop = stop
        last_candle_time = buy_time

    pnl = (sell_price - buy_price) * qty

    # Normalize sell_time to string for consistent Excel output
    st = ""
    try:
        if isinstance(sell_time, time):
            st = sell_time.strftime("%H:%M:%S")
        elif isinstance(sell_time, str):
            st = sell_time
        elif pd.isna(sell_time) or sell_time is None:
            st = ""
        else:
            st = str(sell_time)
    except Exception:
        st = ""
    
    # Normalize last_candle_time to string
    lct = ""
    try:
        if isinstance(last_candle_time, time):
            lct = last_candle_time.strftime("%H:%M:%S")
        elif pd.isna(last_candle_time) or last_candle_time is None:
            lct = ""
        else:
            lct = str(last_candle_time)
    except Exception:
        lct = ""

    # Append to result lists (SINGLE append point)
    sell_prices.append(sell_price)
    sell_times.append(st)
    exit_reasons.append(exit_reason)
    pnls.append(pnl)
    
    # Execution prices (SINGLE append point)
    exec_entry_times.append(buy_time)
    exec_entry_prices.append(buy_price)
    
    # State for Phase-4B continuity (SINGLE append point)
    state_highest_prices.append(highest_price)
    state_current_stops.append(current_stop)
    state_last_candle_times.append(lct)
    state_trailing_armed.append(trailing_armed)

# Assign accumulated lists to DataFrame
df["exec_entry_time"] = exec_entry_times
df["exec_entry_price"] = exec_entry_prices
df["highest_price_state"] = state_highest_prices
df["current_stop_state"] = state_current_stops  
df["last_candle_time_state"] = state_last_candle_times
df["trailing_armed_state"] = state_trailing_armed
df["sell_price"] = sell_prices
df["sell_time"] = sell_times
df["exit_reason"] = exit_reasons
df["pnl"] = pnls


# ===============================
# DETECT OPEN TRADES (LIVE MODE)
# ===============================

if df.empty:
    closed_trades = df.copy()
    open_trades = df.copy()
    logger.info("[Trade Status]")
    logger.info("  Closed: 0")
    logger.info("  Still Open: 0")
else:
    # Create entry datetime for sanity check
    now = datetime.now()
    df["entry_dt"] = pd.to_datetime(
        df["date"].astype(str) + " " + df["exec_entry_time"].astype(str),
        errors="coerce"
    )

    # Only track trades that have already entered (prevent future signals)
    df = df[df["entry_dt"] <= now].copy()

    # Mark trades still open after CSV exhaustion
    df["is_open"] = df["exit_reason"].isin([
        "NO_EXIT",
        "NO_EXIT_LASTCANDLE"
    ])

    # Split into closed and open trades
    closed_trades = df[~df["is_open"]].copy()
    open_trades = df[df["is_open"]].copy()

    logger.info("[Trade Status]")
    logger.info("  Closed: %s", len(closed_trades))
    logger.info("  Still Open: %s", len(open_trades))

    # Save open trades state for Phase-4B
    if len(open_trades) > 0:
        state_file = os.path.join(OUTPUT_DIR, "phase-4b-tracking-state.csv")
        open_trades.to_csv(state_file, index=False)
        logger.info("[SAVED] Open trades saved to: %s", state_file)
        logger.info("[STARTING] Triggering Phase-4B Live Tracker...")
        
        import subprocess
        try:
            subprocess.run([sys.executable, "phase-4b-live.py"], check=True)
        except subprocess.CalledProcessError as e:
            logger.error("Phase-4B Failed: %s", e)
        except KeyboardInterrupt:
            logger.info("Live tracking stopped by user.")

# Use closed_trades for Excel output
df = closed_trades

# Add final columns to df (closed trades only)
df["sell_price"] = sell_prices[:len(df)]
df["sell_time"] = sell_times[:len(df)]
df["exit_reason"] = exit_reasons[:len(df)]
df["pnl"] = pnls[:len(df)]

# Ensure integer quantities
if not df.empty and "quantity" in df.columns:
    df["quantity"] = df["quantity"].astype(int)

# ===============================
# FINAL OUTPUT: Build 4 sheets
# Sheets: Trade Log, Daily Summary, Performance, Algorithm Config
# ===============================

# Prepare Trade Log with requested columns (use existing df columns where available)
trade_log_cols = [
    'Date', 'Stock', 'Mode', 'Weight', 
    'Open', 'High', 'Low', 'Close', 'Volume', 'DataSource', 'CandleTimeframe',
    'EntryTime', 'EntryPrice', 'StopLoss', 'Target', 'ExitTime',
    'ExitReason', 'ExitPrice', 'Quantity', 'InvestedAmount', 'ProfitBeforeCosts',
    'TransactionCost', 'FinalProfit', 'P&L%'
]

# Build trade log rows
trade_rows = []
for _, r in df.iterrows():
    date_val = r.get('date')
    try:
        date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
    except Exception:
        date_str = str(date_val)

    invested = float(r.get('trade_value', 0.0))
    profit_before = float(r.get('pnl', 0.0))
    tx_cost = invested * TRANSACTION_COST_PCT
    final_profit = profit_before - tx_cost
    pl_pct = (final_profit / invested * 100.0) if invested > 0 else 0.0

    # OHLCV will be filled from daily merge file (not from minute data here)
    day_open = np.nan
    day_high = np.nan
    day_low = np.nan
    day_close = np.nan
    day_vol = np.nan

    row = {
        'Date': date_str,
        'Stock': r.get('symbol'),
        'Mode': r.get('mode', ''),
        'Weight': float(r.get('weight', 0.0)),
        'Open': day_open,   # Will be overwritten by daily merge if available
        'High': day_high,
        'Low': day_low,
        'Close': day_close,
        'Volume': day_vol,
        'DataSource': r.get('DataSource') or ONE_MIN_DATA_DIR,
        'CandleTimeframe': r.get('CandleTimeframe') or '1m',
        'EntryTime': r.get('exec_entry_time', r.get('buy_time')),
        'EntryPrice': float(r.get('exec_entry_price', r.get('entry_price', 0.0))),
        'StopLoss': float(r.get('stop_price', np.nan)),
        'Target': float(r.get('target_price', np.nan)),
        'ExitTime': r.get('sell_time'),
        'ExitReason': r.get('exit_reason'),
        'ExitPrice': float(r.get('sell_price', 0.0)),
        'Quantity': int(r.get('quantity', 0)),
        'InvestedAmount': invested,
        'ProfitBeforeCosts': profit_before,
        'TransactionCost': tx_cost,
        'FinalProfit': final_profit,
        'P&L%': pl_pct
    }
    trade_rows.append(row)

trade_log_df = pd.DataFrame(trade_rows, columns=trade_log_cols)

ohlcv_path = os.path.join('downloaded_data', 'daily_candles_nifty500.xlsx')
if os.path.exists(ohlcv_path):
    try:
        ohl = pd.read_excel(ohlcv_path)
        ohl.columns = [str(c).strip() for c in ohl.columns]
        
        # Normalize symbol column
        if 'Symbol' in ohl.columns:
            ohl['Symbol'] = ohl['Symbol'].astype(str).str.upper().str.strip()
        elif 'symbol' in ohl.columns:
            ohl['Symbol'] = ohl['symbol'].astype(str).str.upper().str.strip()
        
        # Handle Datetime column (not Date)
        if 'Datetime' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['Datetime']).dt.strftime('%Y-%m-%d')
        elif 'Date' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['Date']).dt.strftime('%Y-%m-%d')
        elif 'date' in ohl.columns:
            ohl['Date'] = pd.to_datetime(ohl['date']).dt.strftime('%Y-%m-%d')

        # Normalize OHLCV columns to capitalized
        col_map = {
            'open': 'Open', 'high': 'High', 'low': 'Low', 'close': 'Close', 'volume': 'Volume',
            'Open': 'Open', 'High': 'High', 'Low': 'Low', 'Close': 'Close', 'Volume': 'Volume'
        }
        ohl.rename(columns=col_map, inplace=True)

        # keep only relevant OHLCV cols that exist
        ohlv_cols = [c for c in ['Open','High','Low','Close','Volume'] if c in ohl.columns]

        # create lookup dict keyed by (symbol,date)
        ohl['__key'] = ohl['Symbol'].astype(str) + '||' + ohl['Date'].astype(str)
        ohl_lookup = ohl.set_index('__key')[ohlv_cols].to_dict(orient='index')

        # normalize trade keys and attempt fast lookup
        def try_get_ohlv(stock, date):
            if pd.isna(stock) or pd.isna(date):
                return None
            k = str(stock).upper().strip() + '||' + str(date)
            if k in ohl_lookup:
                return ohl_lookup[k]
            # fallback: try exact match in symbols for that date
            for idx,row in ohl[ohl['Date']==str(date)].iterrows():
                sym = row['Symbol']
                if sym == str(stock).upper().strip():
                    return row[ohlv_cols].to_dict()
            # try partial match
            for idx,row in ohl[ohl['Date']==str(date)].iterrows():
                sym = row['Symbol']
                if str(stock).upper().strip() in sym or sym in str(stock).upper().strip():
                    return row[ohlv_cols].to_dict()
            return None

        # apply per-row fill
        for i, tr in trade_log_df.iterrows():
            stock = tr['Stock']
            datev = tr['Date']
            ovals = try_get_ohlv(stock, datev)
            if ovals:
                for c in ohlv_cols:
                    trade_log_df.at[i, c] = ovals.get(c)
        
        # Ensure OHLCV columns are numeric (not strings)
        for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
            if col in trade_log_df.columns:
                trade_log_df[col] = pd.to_numeric(trade_log_df[col], errors='coerce')
                
    except Exception as e:
        logger.warning("Could not load OHLCV data: %s", e)
        pass

# Ensure Bias is readable (only if column exists)
if 'Bias' in trade_log_df.columns:
    trade_log_df['Bias'] = trade_log_df['Bias'].fillna('N/A')

# Daily summary (requested columns)
daily = trade_log_df.groupby('Date').agg(
    TotalTrades=('Stock', 'count'),
    CapitalStart=('InvestedAmount', lambda x: float(C_PER_DAY)),
    CapitalInvested=('InvestedAmount', 'sum'),
    CapitalEnd=('FinalProfit', lambda x: float(C_PER_DAY) + x.sum()),
    Profit=('FinalProfit', lambda x: x[x>0].sum()),
    Loss=('FinalProfit', lambda x: abs(x[x<0].sum())),
    DailyP_and_L=('FinalProfit', 'sum')
).reset_index()

# Rename DailyP_and_L -> DailyP&L and DayStatus
daily = daily.rename(columns={'DailyP_and_L': 'DailyP&L'})
daily['DayStatus'] = daily['DailyP&L'].apply(lambda v: 'PROFIT' if v>0 else ('LOSS' if v<0 else 'BREAKEVEN'))

# Reorder daily columns as requested
daily_summary_cols = ['Date','TotalTrades','CapitalStart','CapitalInvested','CapitalEnd','Profit','Loss','DailyP&L','DayStatus']
daily_summary = daily[daily_summary_cols]

# ===============================
# PERFORMANCE SHEET - CRYSTAL CLEAR FORMAT
# ===============================

# Calculate overall metrics first
total_trades = len(trade_log_df)
wins_df = trade_log_df[trade_log_df['FinalProfit']>0]
losses_df = trade_log_df[trade_log_df['FinalProfit']<0]
win_count = len(wins_df)
loss_count = len(losses_df)
breakeven_count = total_trades - win_count - loss_count

total_invested = trade_log_df['InvestedAmount'].sum()
total_profit_amount = wins_df['FinalProfit'].sum() if win_count > 0 else 0.0
total_loss_amount = abs(losses_df['FinalProfit'].sum()) if loss_count > 0 else 0.0
net_pnl = trade_log_df['FinalProfit'].sum()
roi_percent = (net_pnl / total_invested * 100.0) if total_invested > 0 else 0.0

win_rate = (win_count / total_trades * 100.0) if total_trades > 0 else 0.0
avg_win = wins_df['FinalProfit'].mean() if win_count > 0 else 0.0
avg_loss = abs(losses_df['FinalProfit'].mean()) if loss_count > 0 else 0.0

# Create SUMMARY section (top of sheet)
summary_data = [
    ['═══════════════════════════════════════════════════════════════', ''],
    ['                    📊 BACKTEST PERFORMANCE SUMMARY', ''],
    ['═══════════════════════════════════════════════════════════════', ''],
    ['', ''],
    ['💰 FINAL RESULT', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Capital Deployed', f'₹{total_invested:,.2f}'],
    ['Net Profit/Loss', f'₹{net_pnl:,.2f}'],
    ['Return on Investment (ROI)', f'{roi_percent:.2f}%'],
    ['Final Status', 'PROFIT ✅' if net_pnl > 0 else ('LOSS ❌' if net_pnl < 0 else 'BREAKEVEN')],
    ['', ''],
    ['📈 PROFIT BREAKDOWN', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Profit (from wins)', f'₹{total_profit_amount:,.2f}'],
    ['Total Loss (from losses)', f'₹{total_loss_amount:,.2f}'],
    ['Net Amount', f'₹{net_pnl:,.2f}'],
    ['', ''],
    ['📊 TRADE STATISTICS', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Total Trades', f'{total_trades}'],
    ['Winning Trades', f'{win_count} ({win_rate:.1f}%)'],
    ['Losing Trades', f'{loss_count} ({100-win_rate:.1f}%)'],
    ['Breakeven Trades', f'{breakeven_count}'],
    ['', ''],
    ['💵 AVERAGE PER TRADE', ''],
    ['─────────────────────────────────────────────────────────────────', ''],
    ['Average Win', f'₹{avg_win:,.2f}'],
    ['Average Loss', f'₹{avg_loss:,.2f}'],
    ['Average P&L per Trade', f'₹{net_pnl/total_trades:,.2f}' if total_trades > 0 else '₹0.00'],
    ['', ''],
    ['═══════════════════════════════════════════════════════════════', ''],
]

summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])

# Use summary as the performance sheet (no PromptType breakdown needed)
performance_df = summary_df

# Algorithm Config sheet - produce key/value table from variables
alg_config = {
    'StopLossPercent': f"{S_PCT*100:.2f}%",
    'TargetPercent': f"{(R_MULT*S_PCT)*100:.2f}%",
    'RiskReward': f"1:{R_MULT}",
    'CapitalPerDay': f"₹{C_PER_DAY:,}",
    'CapitalPerTrade': f"₹{int(C_PER_DAY*C_PCT):,}",
    'MaxPositions': int(np.floor(C_PER_DAY / (C_PER_DAY*C_PCT))) if C_PCT>0 else 0,
    'CandleTimeframe': '1m',
    'DataSource': ONE_MIN_DATA_DIR,
    'NewsSource': df.get('NewsSource').iloc[0] if 'NewsSource' in df.columns and len(df) > 0 else '',
    'AIModel': df.get('AIModel').iloc[0] if 'AIModel' in df.columns and len(df) > 0 else '',
    'EntryStartTime': df['buy_time'].iloc[0] if 'buy_time' in df.columns and len(df) > 0 else '',
    'ForceExitTime': FORCE_EXIT_TIME.strftime('%H:%M'),
    'SlippagePercent': f"{SLIPPAGE_PCT*100:.2f}%",
    'TransactionCostPercent': f"{TRANSACTION_COST_PCT*100:.2f}%"
}
config_df = pd.DataFrame(list(alg_config.items()), columns=['Parameter','Value'])

# Write sheets
if len(df) > 0:
    # Write sheets
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        trade_log_df.to_excel(writer, sheet_name='Trade Log', index=False)
        daily_summary.to_excel(writer, sheet_name='Daily Summary', index=False)
        performance_df.to_excel(writer, sheet_name='Performance', index=False)
        config_df.to_excel(writer, sheet_name='Algorithm Config', index=False)

    # Load workbook and apply formatting
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    def format_sheet(ws, header_fill=None, pnl_col_name=None):
        hdr = Font(bold=True)
        for cell in ws[1]:
            cell.font = hdr
            cell.alignment = Alignment(horizontal='center')
            if header_fill:
                cell.fill = header_fill
        if pnl_col_name:
            pnl_col = None
            for idx, cell in enumerate(ws[1], start=1):
                if str(cell.value).lower() == pnl_col_name.lower():
                    pnl_col = idx
                    break
            if pnl_col and ws.max_row > 1:
                col_letter = openpyxl.utils.get_column_letter(pnl_col)
                green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                range_str = f'{col_letter}2:{col_letter}{ws.max_row}'
                ws.conditional_formatting.add(range_str,
                                            CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green))
                ws.conditional_formatting.add(range_str,
                                            CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red))

    # Format Trade Log with blue header
    if 'Trade Log' in wb.sheetnames:
        ws = wb['Trade Log']
        blue = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        format_sheet(ws, header_fill=blue, pnl_col_name='FinalProfit')

    # Format Daily Summary with bold headers and conditional coloring on DailyP&L
    if 'Daily Summary' in wb.sheetnames:
        ws2 = wb['Daily Summary']
        yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        format_sheet(ws2, header_fill=yellow, pnl_col_name='DailyP&L')

    # Format Performance (special handling for summary format)
    if 'Performance' in wb.sheetnames:
        ws3 = wb['Performance']
        
        # Style for section headers (lines with emojis and titles)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, size=14, color='FFFFFF')
        
        # Style for separator lines
        separator_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Style for metric names
        metric_font = Font(bold=True, size=11)
        
        # Style for values
        value_font = Font(size=11)
        
        # Apply formatting row by row
        for row_idx in range(1, ws3.max_row + 1):
            metric_cell = ws3.cell(row=row_idx, column=1)
            value_cell = ws3.cell(row=row_idx, column=2)
            
            metric_text = str(metric_cell.value) if metric_cell.value else ''
            value_text = str(value_cell.value) if value_cell.value else ''
            
            # Header rows (with emojis or equals signs)
            if '═' in metric_text or '📊' in metric_text:
                metric_cell.font = header_font
                metric_cell.fill = header_fill
                value_cell.fill = header_fill
                # Merge cells for header rows
                if '📊' in metric_text or '═' in metric_text:
                    ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                    metric_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Section titles (with emojis)
            elif any(emoji in metric_text for emoji in ['💰', '📈', '📊', '💵']):
                metric_cell.font = Font(bold=True, size=12, color='1F4E78')
                metric_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                value_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            
            # Separator lines
            elif '─' in metric_text:
                metric_cell.fill = separator_fill
                value_cell.fill = separator_fill
                ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            
            # Data rows
            elif metric_text and metric_text.strip():
                metric_cell.font = metric_font
                value_cell.font = value_font
                
                # Color code profit/loss values
                if 'Net Profit/Loss' in metric_text or 'Final Status' in metric_text or 'Net Amount' in metric_text:
                    if '✅' in value_text or (value_text.startswith('₹') and float(value_text.replace('₹','').replace(',','')) > 0):
                        value_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        value_cell.font = Font(bold=True, size=12, color='006100')
                    elif '❌' in value_text or (value_text.startswith('₹') and float(value_text.replace('₹','').replace(',','')) < 0):
                        value_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        value_cell.font = Font(bold=True, size=12, color='9C0006')
                
                # Highlight ROI
                if 'ROI' in metric_text:
                    value_cell.font = Font(bold=True, size=12)
        
        # Adjust column widths
        ws3.column_dimensions['A'].width = 35
        ws3.column_dimensions['B'].width = 25

    # Format Algorithm Config (dark header, clear key/value)
    if 'Algorithm Config' in wb.sheetnames:
        ws4 = wb['Algorithm Config']
        dark = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')
        for cell in ws4[1]:
            cell.fill = dark
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
        # alternate shading for readability
        for r in range(2, ws4.max_row+1):
            fill = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid') if r%2==0 else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            for c in range(1, ws4.max_column+1):
                ws4.cell(row=r, column=c).fill = fill

    wb.save(OUTPUT_FILE)

    logger.info("Phase-4 BACKTEST (1-minute) completed successfully")
    logger.info("Saved to: %s", OUTPUT_FILE)
    logger.info("Total P&L: INR %,.2f", df['pnl'].sum())

else:
    logger.info("No closed trades in backtest (all trades live/pending).")
    logger.info("Skipped generating empty backtest report.")

# Always try to update visualization (picks up latest result file)
logger.info("Updating Visualization Data (viz_data.json)...")
try:
    import subprocess
    subprocess.run(["python", "prepare_viz_data.py"], check=True)
    logger.info("Visualization data prepared.")
except Exception as e:
    logger.error("Failed to run prepare_viz_data.py: %s", e)
